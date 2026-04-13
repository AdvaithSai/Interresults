"""
scraper.py  –  TGBIE (Telangana Board of Intermediate Education) Mark Fetcher
Website : https://tgbienew.cgg.gov.in/ResultMemorandum.do

Run:
    python scraper.py

Output:
    marksheet_output.xlsx  — one row per student with all subject marks + total
"""

import sys
import time
import traceback
import re
import pandas as pd
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException, TimeoutException, StaleElementReferenceException
)
from webdriver_manager.chrome import ChromeDriverManager

import config as cfg


# ══════════════════════════════════════════════════════════════════════════════
#  BROWSER SETUP
# ══════════════════════════════════════════════════════════════════════════════

import platform

def setup_driver() -> webdriver.Chrome:
    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("useAutomationExtension", False)
    
    # Cloud / Docker compatibility
    if platform.system() != "Windows":
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        
    driver_path = ChromeDriverManager().install()
    if "THIRD_PARTY_NOTICES" in driver_path:
        driver_path = driver_path.replace("THIRD_PARTY_NOTICES.chromedriver", "chromedriver")
        
    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=options)
    driver.set_window_size(1400, 900)
    return driver


def wait_for(driver, by, value, timeout=10):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, value))
    )


def wait_clickable(driver, by, value, timeout=10):
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, value))
    )


# ══════════════════════════════════════════════════════════════════════════════
#  FORM FILLING
# ══════════════════════════════════════════════════════════════════════════════

def fill_form(driver, hall_ticket: str, exam_year: str | None = None):
    """Fill the TGBIE memo download form and click Get Memo.

    exam_year overrides cfg.EXAM_YEAR when provided
    (pass 'First Year' or 'Second Year').
    """
    year = (exam_year or cfg.EXAM_YEAR).strip().lower()

    # 1. Results Year dropdown
    year_select = wait_for(driver, By.ID, cfg.RESULTS_YEAR_ID)
    sel = Select(year_select)
    try:
        sel.select_by_value(cfg.RESULTS_YEAR)
    except Exception:
        sel.select_by_index(0)

    # 2. Year radio  (First Year → year1 / Second Year → year2)
    if year == "second year":
        driver.find_element(By.ID, cfg.YEAR_RADIO_YEAR2).click()
    else:
        driver.find_element(By.ID, cfg.YEAR_RADIO_YEAR1).click()

    # 3. Category radio
    cat = cfg.CATEGORY.strip().lower()
    if "vocational bridge" in cat:
        driver.find_element(By.ID, "categoryVBC").click()
    elif "general bridge" in cat:
        driver.find_element(By.ID, "categoryGBC").click()
    elif "vocational" in cat:
        driver.find_element(By.ID, cfg.CATEGORY_VOCA).click()
    else:
        driver.find_element(By.ID, cfg.CATEGORY_GENERAL).click()

    # 4. Exam type radio  IPE=3 / IPASE=6
    exam_val = "6" if cfg.EXAM_TYPE.strip().upper() == "IPASE" else "3"
    driver.find_element(
        By.XPATH, f"//input[@name='property(month)' and @value='{exam_val}']"
    ).click()

    # 5. Hall Ticket Number
    ht_input = wait_for(driver, By.ID, cfg.HALL_TICKET_ID)
    ht_input.clear()
    ht_input.send_keys(hall_ticket)

    # 6. Submit button (type="button", triggers validate() JS)
    submit_btn = wait_clickable(driver, By.XPATH,
        "//input[@type='button' and @value='Get Memo']")
    submit_btn.click()


# ══════════════════════════════════════════════════════════════════════════════
#  RESULT EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def clean(text: str) -> str:
    """Remove asterisks (supplementary marker) and extra spaces."""
    return text.replace("*", "").strip()


def extract_all_data(driver) -> dict:
    """
    Extract ALL student data using JS with the EXACT cell structure of TGBIE:

    Info labels:  cell[i] = 'REGD NUMBER\n:'  cell[i+1] = '2658266175'
    Mark cells:   cell = 'Max_val\nSecured_val * P' (both rows stacked due to rowspan)
    PART-III:     ONE cell = 'MathsA_max\nMathsB_max\nPhysics_max\nChemistry_max\n
                              MathsA_sec\nMathsB_sec\nPhysics_sec\nChemistry_sec'
    Grand Total:  cell = 'GRAND TOTAL:   700  GRAND TOTAL: (in words)...'
    Result:       cell = 'RESULT:   B GRADE   DATE: ...'
    """
    js = r"""
    var result = {};
    var allCells = Array.from(document.querySelectorAll('td, th'));
    var cellTexts = allCells.map(c => c.innerText.trim());

    // ── 1. STUDENT INFO ─────────────────────────────────────────────────────
    // Labels are like 'REGD NUMBER\n:' and value is next non-empty cell
    var infoMap = {
        'REGD NUMBER': 'Regd_No',
        'NAME': 'Name'
    };
    // Father/Mother have apostrophe
    for (var i = 0; i < cellTexts.length; i++) {
        var t = cellTexts[i];
        // Normalize: strip \n : and trailing whitespace for comparison
        var key = t.replace(/[\n\r:]/g, '').trim().toUpperCase();
        if (key === 'REGD NUMBER') result['Regd_No'] = cellTexts[i+1] || '';
        if (key === 'NAME') result['Name'] = cellTexts[i+1] || '';
        if (key === "FATHER'S NAME" || key === 'FATHERS NAME') result['Father'] = cellTexts[i+1] || '';
        if (key === "MOTHER'S NAME" || key === 'MOTHERS NAME') result['Mother'] = cellTexts[i+1] || '';
        if (key === 'GENDER') result['Gender'] = (cellTexts[i+1] || '').split(/\s/)[0];
        if (t.startsWith('MEDIUM') && t.includes(':')) {
            var mMatch = t.match(/MEDIUM\s*[:\xa0]\s*(\w+)/i);
            if (mMatch) result['Medium'] = mMatch[1];
        }
    }

    // ── 2. MARKS ─────────────────────────────────────────────────────────────
    // Table structure (each data row is ONE <td> per subject group due to rowspan):
    //   Col pattern for 1st/2nd YEAR row:
    //     [year_cell]  [label_cell='Max Marks\nMarks Secured']  [eng_theory_cell]  [eng_prac_cell]  [sanskrit_cell]  [part3_cell]
    //   part3_cell = 'MathsA_max\nMathsB_max\nPhysics_max\nChemistry_max\nMathsA_sec\nMathsB_sec\nPhysics_sec\nChemistry_sec'
    //   Practicals row:
    //     [PRACTICALS_cell]  [label_cell]  [prac_phy_chem_cell='30\n30\n26 P\n29 P']

    // Find all rows in the marks table that contain 'Marks Secured'
    var tables = document.querySelectorAll('table');
    var marksTable = null;
    for (var t of tables) {
        if (t.innerText.includes('Marks Secured')) { marksTable = t; break; }
    }

    function securedFromCell(cellText) {
        // A cell stacks Max on line 1, Secured on line 2 (possibly with * P or F)
        // e.g. '80\n60*  P'  -> '60'
        //      '80\n09F'     -> '9 FAIL'
        //      '80\n12* F'   -> '12 FAIL'
        var lines = cellText.split('\n').map(s => s.trim()).filter(s => s.length > 0);
        if (lines.length >= 2) {
            var sec = lines[lines.length - 1];
            var numMatch = sec.match(/^(\d+)/);
            if (!numMatch) return sec.includes('AB') ? 'AB' : '';
            var mark = numMatch[1];
            // Detect fail: 'F' appears anywhere after the digits (e.g. '09F', '12* F', '9 F')
            var isFail = /\d[*\s]*F/i.test(sec);
            return isFail ? mark + ' FAIL' : mark;
        }
        return '';
    }

    function securedValuesFromPart3Cell(cellText) {
        // PART-III cell has 4 max marks then 4 secured marks
        // '75\n75\n60\n60\n46* P\n43* P\n34* P\n21* P'  -> all pass
        // '75\n75\n60\n60\n46* P\n43* P\n09F\n21* P'     -> Physics FAIL
        var lines = cellText.split('\n').map(s => s.trim()).filter(s => s.length > 0);
        var half = Math.floor(lines.length / 2);
        var secLines = lines.slice(half);
        return secLines.map(function(s) {
            var numMatch = s.match(/^(\d+)/);
            if (!numMatch) return s.includes('AB') ? 'AB' : '';
            var mark = numMatch[1];
            var isFail = /\d[*\s]*F/i.test(s);
            return isFail ? mark + ' FAIL' : mark;
        });
    }

    if (marksTable) {
        var rows = Array.from(marksTable.querySelectorAll('tr'));

        var headerText = "";
        for (var hrow of rows) {
            var hcells = Array.from(hrow.querySelectorAll('td, th'));
            for (var hc of hcells) {
                headerText += " " + hc.innerText.trim().toUpperCase();
            }
        }
        
        var group = 'MPC';   // default
        if (headerText.includes('BOTANY') || headerText.includes('ZOOLOGY')) {
            group = 'BPC';
        } else if (headerText.includes('HISTORY') || headerText.includes('CIVICS') || headerText.includes('POLITICAL SCIENCE')) {
            group = 'HEC';
        } else if (headerText.includes('ECONOMICS') || headerText.includes('COMMERCE')) {
            group = 'MEC';
        } else if (headerText.includes('MATHEMATICS') || headerText.includes('PHYSICS')) {
            group = 'MPC';
        }
        result['Group'] = group;

        // PART-III subject keys per group
        var p3keys;
        if (group === 'BPC') {
            p3keys = ['Botany', 'Zoology', 'Physics', 'Chemistry'];
        } else if (group === 'HEC') {
            p3keys = ['Economics', 'History', 'Civics'];
        } else if (group === 'MEC') {
            p3keys = ['Maths_A', 'Maths_B', 'Economics', 'Commerce'];
        } else {
            p3keys = ['Maths_A', 'Maths_B', 'Physics', 'Chemistry'];  // MPC
        }

        for (var row of rows) {
            var cells = Array.from(row.querySelectorAll('td'));
            if (cells.length === 0) continue;

            var prefix = null;
            var labelCellIdx = -1;

            for (var ci = 0; ci < cells.length; ci++) {
                var ct = cells[ci].innerText.trim().toUpperCase();
                if (ct.includes('1ST') && ct.includes('YEAR')) prefix = 'Y1';
                else if (ct.includes('2ND') && ct.includes('YEAR')) prefix = 'Y2';
                else if (ct.includes('PRACTICAL')) prefix = 'PR';
                if (cells[ci].innerText.trim().includes('Max Marks')) {
                    labelCellIdx = ci;
                }
            }

            if (!prefix || labelCellIdx < 0) continue;

            if (prefix === 'PR') {
                var pracCell = cells[labelCellIdx + 1];
                if (pracCell) {
                    var lines = pracCell.innerText.trim().split('\n').map(s => s.trim()).filter(s => s);
                    var half = Math.floor(lines.length / 2);
                    var phyMatch = lines[half]     ? lines[half].match(/^(\d+)/)   : null;
                    var cheMatch = lines[half + 1] ? lines[half+1].match(/^(\d+)/) : null;
                    if (phyMatch) result['PR_Physics']   = phyMatch[1];
                    if (cheMatch) result['PR_Chemistry'] = cheMatch[1];
                }
                continue;
            }

            var markStart = labelCellIdx + 1;
            var subjectCells = cells.slice(markStart);

            if (subjectCells.length >= 1)
                result[prefix + '_Eng_Theory'] = securedFromCell(subjectCells[0].innerText.trim());
            if (subjectCells.length >= 2)
                result[prefix + '_Eng_Prac'] = securedFromCell(subjectCells[1].innerText.trim());
            if (subjectCells.length >= 3)
                result[prefix + '_Sanskrit'] = securedFromCell(subjectCells[2].innerText.trim());
            if (subjectCells.length >= 4) {
                var p3vals = securedValuesFromPart3Cell(subjectCells[3].innerText.trim());
                for (var pi = 0; pi < p3keys.length && pi < p3vals.length; pi++) {
                    result[prefix + '_' + p3keys[pi]] = p3vals[pi];
                }
            }
        }
    }

    // ── 3. GRAND TOTAL & RESULT ───────────────────────────────────────────────
    for (var i = 0; i < cellTexts.length; i++) {
        var ct = cellTexts[i];
        if (ct.includes('GRAND TOTAL:')) {
            var gm = ct.match(/GRAND TOTAL:\s*([\xa0\s]*)?(\d+)/);
            if (gm) result['Grand_Total'] = gm[2];
        }
        if (ct.includes('RESULT:')) {
            var rm = ct.match(/RESULT:\s*([\xa0\s]*)?([A-Z] GRADE)/);
            if (rm) result['Result'] = rm[2].trim();
        }
    }

    return result;
    """

    try:
        raw = driver.execute_script(js)
    except Exception as e:
        print(f"    JS extraction error: {e}")
        return {}

    return raw or {}


# Keep old helpers as no-ops 
def extract_student_info(driver): return {}
def extract_marks_table(driver): return {}
def extract_totals(driver): return {}


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN LOOP
# ══════════════════════════════════════════════════════════════════════════════

def fetch_one(driver, hall_ticket: str, exam_year: str | None = None) -> dict | None:
    try:
        driver.get(cfg.WEBSITE_URL)
        time.sleep(1.5)

        fill_form(driver, hall_ticket, exam_year=exam_year)
        time.sleep(cfg.PAGE_LOAD_WAIT)

        # Check if result is present (quick sanity check on page title/content)
        page_title = driver.title.lower()
        if "memo download" not in page_title and "memorandum" not in page_title:
            # Try body check as fallback
            try:
                body_text = driver.find_element(By.TAG_NAME, "body").text[:500]
                if "invalid" in body_text.lower() or "not found" in body_text.lower():
                    print(f"    No result for {hall_ticket}")
                    return None
            except Exception:
                pass

        # Extract everything in one JS call
        data = extract_all_data(driver)

        if not data:
            print(f"    No data extracted for {hall_ticket}")
            return None

        student = {"Hall_Ticket": hall_ticket}
        student.update(data)

        name = data.get('Name', '?')
        gt   = data.get('Grand_Total', '?')
        print(f"    {hall_ticket}  {name:<25}  Total: {gt}")
        return student

    except TimeoutException:
        print(f"    Timeout: {hall_ticket}")
    except Exception as e:
        print(f"    Error [{hall_ticket}]: {e}")
        if "--debug" in sys.argv:
            traceback.print_exc()
    return None


def _style_sheet(ws, df_cols):
    """Apply consistent styling to a worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill("solid", fgColor="1F4E79")
    fail_fill   = PatternFill("solid", fgColor="FF4444")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    # Header row
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = thin_border

    # Data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill_color = "EBF3FB" if row_idx % 2 == 0 else "FFFFFF"
        row_fill   = PatternFill("solid", fgColor=fill_color)
        for cell in row:
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal="center")
            if isinstance(cell.value, str) and "FAIL" in cell.value:
                cell.fill = fail_fill
                cell.font = Font(bold=True, color="FFFFFF", size=10)
            else:
                cell.fill = row_fill

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 35)

    # Freeze header
    ws.freeze_panes = "A2"


def save_excel(data: list[dict], path: str):
    df = pd.DataFrame(data)
    df.drop(columns=["_exam_year"], errors="ignore", inplace=True)  # internal tag

    # ── Column rename maps ────────────────────────────────────────────────────
    common_rename = {
        "Hall_Ticket": "Hall Ticket No",
        "Regd_No":     "Regd No",
        "Name":        "Name",
        "Father":      "Father's Name",
        "Mother":      "Mother's Name",
        "Gender":      "Gender",
        "Medium":      "Medium",
        "Y1_Eng_Theory": "1Y English (Theory)",
        "Y1_Eng_Prac":   "1Y English (Prac)",
        "Y1_Sanskrit":   "1Y 2nd Language",
        "Y2_Eng_Theory": "2Y English (Theory)",
        "Y2_Eng_Prac":   "2Y English (Prac)",
        "Y2_Sanskrit":   "2Y 2nd Language",
        "PR_Physics":    "Practical Physics",
        "PR_Chemistry":  "Practical Chemistry",
        "Grand_Total":   "Grand Total",
        "Result":        "Result Grade",
    }

    mpc_rename = {
        **common_rename,
        "Y1_Maths_A":  "1Y Maths A",    "Y1_Maths_B":   "1Y Maths B",
        "Y1_Physics":  "1Y Physics",    "Y1_Chemistry": "1Y Chemistry",
        "Y2_Maths_A":  "2Y Maths A",    "Y2_Maths_B":   "2Y Maths B",
        "Y2_Physics":  "2Y Physics",    "Y2_Chemistry": "2Y Chemistry",
    }

    bpc_rename = {
        **common_rename,
        "Y1_Botany":   "1Y Botany",     "Y1_Zoology":   "1Y Zoology",
        "Y1_Physics":  "1Y Physics",    "Y1_Chemistry": "1Y Chemistry",
        "Y2_Botany":   "2Y Botany",     "Y2_Zoology":   "2Y Zoology",
        "Y2_Physics":  "2Y Physics",    "Y2_Chemistry": "2Y Chemistry",
    }

    mec_rename = {
        **common_rename,
        "Y1_Maths_A":  "1Y Maths A",    "Y1_Maths_B":   "1Y Maths B",
        "Y1_Economics": "1Y Economics", "Y1_Commerce":  "1Y Commerce",
        "Y2_Maths_A":  "2Y Maths A",    "Y2_Maths_B":   "2Y Maths B",
        "Y2_Economics": "2Y Economics", "Y2_Commerce":  "2Y Commerce",
    }

    hec_rename = {
        **common_rename,
        "Y1_Economics": "1Y Economics", "Y1_History":   "1Y History", "Y1_Civics":    "1Y Civics",
        "Y2_Economics": "2Y Economics", "Y2_History":   "2Y History", "Y2_Civics":    "2Y Civics",
    }

    # ── Column order helpers ──────────────────────────────────────────────────
    first_cols = ["Hall Ticket No", "Regd No", "Name",
                  "Father's Name", "Mother's Name", "Gender", "Medium"]
    last_cols  = ["Grand Total", "Result Grade"]

    def ordered_cols(df_renamed):
        mid = [c for c in df_renamed.columns if c not in first_cols + last_cols]
        return ([c for c in first_cols if c in df_renamed.columns]
                + mid
                + [c for c in last_cols  if c in df_renamed.columns])

    # ── Split into MPC / BPC / MEC / HEC ───────────────────────────────────────────────
    grp = df.get("Group", pd.Series(["MPC"] * len(df)))
    mpc_data = df[grp == "MPC"].copy()
    bpc_data = df[grp == "BPC"].copy()
    mec_data = df[grp == "MEC"].copy()
    hec_data = df[grp == "HEC"].copy()

    output_path = Path(path)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        for sheet_name, sheet_df, rename_map in [
            ("MPC", mpc_data, mpc_rename),
            ("BPC", bpc_data, bpc_rename),
            ("MEC", mec_data, mec_rename),
            ("HEC", hec_data, hec_rename),
        ]:
            if sheet_df.empty:
                continue

            out = sheet_df.drop(columns=["Group"], errors="ignore")
            out = out.rename(columns=rename_map)
            out = out[[c for c in ordered_cols(out) if c in out.columns]]

            out.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            _style_sheet(ws, list(out.columns))

    counts = f"{len(mpc_data)} MPC + {len(bpc_data)} BPC + {len(mec_data)} MEC"
    print(f"\n+  Saved {counts} students  ->  {output_path.resolve()}\n")


def run_batch(driver, roll_numbers: list, exam_year: str, label: str) -> list:
    """Fetch marksheets for a list of roll numbers under a given exam year."""
    data, failed = [], []
    total = len(roll_numbers)

    print(f"\n{'='*65}")
    print(f"  BATCH : {label}  ({exam_year})")
    print(f"  Students : {total}")
    print(f"{'='*65}\n")

    for i, ht in enumerate(roll_numbers, 1):
        print(f"[{i:>3}/{total}]  {ht}", end="  ")
        result = fetch_one(driver, ht, exam_year=exam_year)
        if result:
            result["_exam_year"] = exam_year   # tag for filtering
            data.append(result)
        else:
            failed.append(ht)
        if i < total:
            time.sleep(1)

    if failed:
        tag = "1Y" if "first" in exam_year.lower() else "2Y"
        fname = f"failed_{tag}.txt"
        Path(fname).write_text("\n".join(failed))
        print(f"\n  ❌ {len(failed)} failed → saved to {fname}")

    return data


def main():
    print("\n" + "=" * 65)
    print("  TGBIE MARKS AUTOMATION")
    print(f"  Website  : {cfg.WEBSITE_URL}")
    print(f"  Category : {cfg.CATEGORY}  |  Type: {cfg.EXAM_TYPE}  |  Year: {cfg.RESULTS_YEAR}")
    print(f"  2nd Year : {len(cfg.ROLL_NUMBERS)} students  →  {cfg.OUTPUT_FILE}")
    fy_count = len(getattr(cfg, 'FIRST_YEAR_ROLL_NUMBERS', []))
    if fy_count:
        print(f"  1st Year : {fy_count} students  →  {cfg.FIRST_YEAR_OUTPUT_FILE}")
    print("=" * 65)

    driver = setup_driver()

    try:
        # ── 2nd Year batch ────────────────────────────────────────────────────
        if cfg.ROLL_NUMBERS:
            data_2y = run_batch(driver, cfg.ROLL_NUMBERS,
                                exam_year="Second Year", label="2nd Year")
            if data_2y:
                save_excel(data_2y, cfg.OUTPUT_FILE)

        # ── 1st Year batch ────────────────────────────────────────────────────
        fy_rolls = getattr(cfg, 'FIRST_YEAR_ROLL_NUMBERS', [])
        if fy_rolls:
            data_1y = run_batch(driver, fy_rolls,
                                exam_year="First Year", label="1st Year")
            if data_1y:
                save_excel(data_1y, cfg.FIRST_YEAR_OUTPUT_FILE)

    except KeyboardInterrupt:
        print("\n\n⚠  Interrupted — partial data may have been saved.\n")
    finally:
        driver.quit()

    print("\n🎉  Done!")


if __name__ == "__main__":
    main()
